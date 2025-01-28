"""Парсер расписания с сайта vega.mirea.ru"""

# std
from os import getenv
from os import path

from re import fullmatch
from re import search
from copy import deepcopy
# datetime
from datetime import datetime
from datetime import time
from datetime import timedelta

# usefull
from openpyxl import load_workbook
from progress.bar import PixelBar
from dotenv import load_dotenv

# custom
from src import utils
from src.db_class import Database

WEEKS = 17
AUTUMN_SUBSTR = "осен"
SPRING_SUBSTR = "весен"
WINTER_SUBSTR = "зимн"
SUMMER_SUBSTR = "летн"


# TODO распараллелить
# TODO прогресс бар как декоратор
# TODO брать расписание с сайта ВЕГИ
# TODO сделать НИР костылем, в котором 

class VegaRaspParser:
    """Парсер excel расписания"""

    def __init__(self) -> None:
        self.db = Database()
        # тип файла для парсинга
        load_dotenv()
        self.default_filename = getenv("DEFAULT_FILENAME")
        self.session_filename = getenv("SESSION_FILENAME")

        self.is_default = False
        self.is_session = False

        if self.default_filename is None and self.session_filename is None:
            raise ValueError("Не выбран ни один файл для парсинга")

        if self.default_filename is not None:
            if not path.exists(self.default_filename):
                raise ValueError(
                    f"Выбран несуществующий файл файл расписания семестра: {self.default_filename}")
            self.is_default = True

        if self.session_filename is not None:
            if not path.exists(self.session_filename):
                raise ValueError(
                    f"Выбран несуществующий файл расписания сессии: {self.session_filename}")
            self.is_session = True

        self.group_row = 2
        self.weekday_col = 1
        self.para_col = 2

        # позиция ячейки с версией
        self.version_col = 1
        self.version = 0

        self.subgroups = ["(1пг)", "(2пг)"]
        self.parity = ["Iн", "IIн"]
        self.week_strs = ["ВС", "ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ"]

        # дата начала и конца семестра
        self.start_date = getenv("START_DATE")
        self.end_date = getenv("END_DATE")
        if self.start_date is None or self.end_date is None:
            raise ValueError("Добавьте даты начала и конца семестра")

        # код семестра расписания
        self.semcode = 0

        # Год начала и конца обучения
        self.start_year = datetime.now().year
        self.end_year = self.start_year + 1

    def change_to_session_params(self):
        """Меняет некоторые стандартные параметры на параметры расписания сессии"""
        self.version_col = 2
        self.group_row = 4

    def change_to_dafault_params(self):
        """Возвращает параметры к дефолтным"""
        # позиция ячейки с версией
        self.version_col = 1
        self.group_row = 2

    def parse(self):
        """Парсинг исходя из параметров .env"""
        self.db.set_conn()

        # заполняем кафедры
        self.db.fill_departments()

        # заполняем типы пар
        self.db.fill_worktypes()

        # если выбрали файл расписания семетра
        if self.is_default:
            print("<--PARSING SEMESTR-->")
            self.change_to_dafault_params()
            self.parse_semestr()
            print("<--RASP PARSED-->")
        # если выбрали файл расписания сессии
        if self.is_session:
            print("<--PARSING SESSION-->")
            self.change_to_session_params()
            self.parse_exam()
            print("<--SESSION PARSED-->")

        self.db.close_conn()

    def parse_semestr(self):
        """Парсинг расписания семестровых пар"""
        self.parse_excel_file(self.default_filename)

    def parse_exam(self):
        """Парсинг расписания экзаменов"""
        self.parse_excel_file(self.session_filename, False)

    def parse_excel_file(self, file_name, default_rasp=True):
        """Парсим данные"""
        ws = load_workbook(filename=file_name).active
        # ------------------Определяем границы расписания------------------
        min_col = ws.min_column
        max_col = ws.max_column
        max_row = self.get_max_row(ws)
        # -----------------------------------------------------------------
        # вытаскиваем все нужное из заголовка
        self.fill_rasp_title_parts(max_row, ws)

        # заполняем дни таблицы rasp18_days
        self.db.fill_rasp18_for_period(
            self.semcode, self.start_date, self.end_date)
        # ---------------Парсим расписание в цикле по группам--------------
        for col in range(min_col, max_col + 1):
            group_name = ws.cell(self.group_row, col).value
            # если заголовок столбца пустой, пропускаем
            if not group_name:
                continue
            merged_cells = ws.merged_cells
            if default_rasp:
                self.parse_default_col(
                    col, ws, max_row, group_name, merged_cells
                )
            else:
                self.parse_exam_col(
                    col, ws, max_row, group_name
                )

    def get_max_row(self, ws):
        """Находим последнюю значимую строку"""
        max_row = ws.max_row
        max_col = ws.max_column
        legend_pattern = r"(\w)*егенд[\w\s]*"

        for row in range(max_row, 1, -1):
            for col in range(max_col, 1, -1):
                cur_cell = ws.cell(row, col).value
                if cur_cell is None:
                    continue
                # если нашли строку со словом "легенда", то далее ищем первую значимую строку
                if fullmatch(legend_pattern, ws.cell(row, col).value):
                    max_row = row
                    break
            if max_row != ws.max_row:
                break

        for row in range(max_row - 1, 1, -1):
            if not utils.is_hsplitter(ws, row):
                return row
        return max_row

    def parse_default_col(self, col, ws, max_row, group_name, merged_cells):
        """Разбор колонки группы"""
        # заполняем БД данными по группе
        group_id = self.db.set_group(group_name)

        order = 1
        prev_weekday = "ПН"

        # проходимся по текущему столбцу по всем строкам
        row = self.group_row + 1

        progress_bar = PixelBar(
            group_name, max=max_row-row, suffix='%(percent)d%%')
        progress_bar.check_tty = False
        progress_bar.start()

        while row < max_row:
            progress_bar.next()

            weekday = ws.cell(row, self.weekday_col).value
            # если новый день недели
            if prev_weekday != weekday and weekday:
                prev_weekday = weekday
            weekday = prev_weekday

            # вытаскиваем данные по текущей паре
            cur_order = ws.cell(row, self.para_col).value
            lesson_cell = ws.cell(row, col).value
            teacher = ws.cell(row, col + 1).value
            room = ws.cell(row, col + 2).value

            # если пустая клетка, идем дальше
            if lesson_cell is None:
                row += 1
                # если строка нечетная, то номер пары + 1
                if row % 2 == 0:
                    order += 1
                continue

            # является ли длинной парой (НИР, ВОЕНКА, условно на весь день)
            coord = ws.cell(row, col).coordinate
            lesson_count = utils.get_lesson_count(merged_cells, coord)

            # по цвету определяем пренадлежность к кафедре
            cur_color = ws.cell(row, col).fill.start_color.rgb
            department_id = self.get_dep_id(cel_color=cur_color)

            # если смешаная клетка
            if cur_order is not None:
                order = cur_order

            lesson_parts = self.get_lesson_parts(lesson_cell)

            # заполнение БД новыми данными
            for pair_num in range(order, order+lesson_count):
                self.fill_group_day_db(
                    lesson_parts,
                    teacher,
                    pair_num,
                    weekday,
                    group_id,
                    room,
                    department_id,
                )

            # переходим к новой паре
            row += 1
        progress_bar.finish()

    def parse_exam_col(self, col, ws, max_row, group_cell):
        """Разбор колонки расписания экзаменов"""
        # убираем курс из названия группы

        # Делит ячейку группы на части: 1 курс\nКМБО-47-25
        _, group_name = group_cell.split("\n")

        # заполняем БД данными по группе
        group_id = self.db.set_group(group_name)

        prev_date_cell = None
        row = self.group_row + 2

        progress_bar = PixelBar(
            group_name, max=max_row-row, suffix='%(percent)d%%')
        progress_bar.check_tty = False
        progress_bar.start()

        # проходимся по текущему столбцу по всем строкам
        prev_weekday = "ПН"
        while row < max_row:
            date_cell = ws.cell(row, self.para_col).value
            # если следующий день
            prev_date_cell, date_cell = utils.swap_with_prev_value(
                prev_date_cell, date_cell
            )
            exam_date, weekday = date_cell.split("\n")
            if prev_weekday != weekday:
                prev_weekday = weekday

            exam_type = ws.cell(row, col + 1).value
            lesson_name = ws.cell(row + 1, col).value
            teacher = ws.cell(row + 2, col).value
            room = ws.cell(row + 2, col + 1).value

            # время начала экзамена
            time_start = ws.cell(row, col).value
            time_end = time_start

            if time_start is None:
                # если только эта строка пустая, то это разделитель
                is_splitter_row = utils.is_hsplitter(ws, row)
                if is_splitter_row:
                    progress_bar.next()
                    row += 1
                    continue

                # если все ячейки данных по экзамену пустые
                if lesson_name is None and teacher is None:
                    progress_bar.next(3)
                    row += 3
                    continue

                # если предмет не пустой, то ставим на 4 пары
                if teacher is None:
                    time_start = time(hour=9, minute=0)
                    time_end = time(hour=15, minute=50)

            # ставим длительность экзамена 1.5 часа (2 академ часа)
            else:
                time_end = utils.time_in_90_minutes(time_start)

            cur_order = utils.get_order_by_time(time_start)

            # Заполение таблиц
            month = int(exam_date[-2:])
            if month > 8:
                exam_date += "." + str(self.start_year)
            else:
                exam_date += "." + str(self.end_year)

            weekday_num = self.week_strs.index(weekday.upper())
            week = self.db.get_week_by_date(exam_date)
            worktype = utils.get_worktype(exam_type)
            cur_color = ws.cell(row, col + 1).fill.start_color.index
            department_id = self.get_dep_id(cel_color=cur_color)
            disc_id = self.get_disc_id(lesson_name, department_id)

            # если перепутали название дисциплины, добавим такую
            if disc_id is None:
                disc_id = self.db.set_disc(
                    title="null",
                    shorttitle=lesson_name,
                    department_id=department_id,
                    varmask="null",
                )

            # если новый день, то добавляем
            rasp18_days_id = self.db.set_rasp18_days(
                semcode=self.semcode, day=exam_date, weekday=weekday_num, week=week
            )
            # set_rasp18
            rasp18_id = self.db.set_rasp18(
                semcode=self.semcode,
                day_id=rasp18_days_id,
                pair=cur_order,
                kind=0,
                worktype=worktype,
                disc_id=disc_id,
                timestart=str(time_start),
                timeend=str(time_end),
            )
            # set_rasp18_groups
            self.db.set_rasp18_groups(
                rasp18_id=rasp18_id, group_id=group_id, subgroup=0
            )
            # set_rasp18_preps
            if teacher is not None:
                params = {"fio": teacher}
                prep_id = self.db.get_id(table_name="sc_prep", params=params)
                # если перепутали имя препода, добавим такую
                if prep_id is None:
                    prep_id = self.db.set_prep(fio=teacher)

                self.db.set_rasp18_preps(rasp18_id=rasp18_id, prep_id=prep_id)
            # set_rasp18_rooms
            if room is not None:
                self.db.set_rasp18_rooms(rasp18_id=rasp18_id, room=room)

            # пропускаем уже разобранные ячейки экзамена
            progress_bar.next(3)
            row += 3

        progress_bar.finish()

    def get_lesson_parts(self, lesson_cell: str) -> dict:
        """
        Разделяет пару на название дисциплины и доп информацию:
        подгруппа, недели, тип пары (лк, пр, лб)
        """
        lesson_parts = {
            "disc_name": "",
            "sub_group": self.get_subgroup(lesson_cell),
            "parity": 0,
            "weeks_list": "null",
            "weeks_text": "null",
            "worktype": "null",
        }

        # получаем тип пары (лк, пр, лб)
        lesson_type = utils.get_lesson_type(lesson_cell)
        worktype = utils.get_worktype(lesson_type)
        # получаем подгруппу
        # получаем данные по неделям
        weeks_parts = self.get_weeks_parts(lesson_cell)
        lesson_parts["parity"] = weeks_parts["parity"]
        lesson_parts["weeks_list"] = weeks_parts["weeks_list"]
        lesson_parts["weeks_text"] = weeks_parts["weeks_text"]
        lesson_parts["worktype"] = worktype

        # удаляем лишнее и получаем название дисциплины
        lesson_parts["disc_name"] = self.get_disc_name(
            lesson_cell, lesson_parts)

        return lesson_parts

    def get_subgroup(self, lesson: str) -> int:
        """Получить подгруппу"""
        for subgroup in self.subgroups:
            if subgroup in lesson:
                return self.subgroups.index(subgroup) + 1
        return 0

    def get_weeks_parts(self, lesson: str) -> dict:
        """Получить номера недель, когда будет пара"""
        all_weeks = list(range(1, WEEKS))
        weeks_text = ", ".join(map(str, all_weeks))

        weeks_parts = {"parity": 0, "weeks_list": all_weeks,
                       "weeks_text": weeks_text}

        parity = self.get_week_parity(lesson)
        weeks_parts["parity"] = parity
        # генирируем массив всех недель
        if parity != 0:
            weeks_parts["weeks_list"] = list(range(parity, WEEKS, 2))
            weeks_parts["weeks_text"] = self.parity[parity - 1]
            return weeks_parts

        pattern = r"\d+(,(\s)?\d+)*н"
        substring = search(pattern, lesson)
        if substring is not None:
            # проверяем наличие недельного распределения
            weeks_parts["weeks_text"] = substring.group()
            substring = substring.group().removesuffix("н")
            weeks = substring.split(",")
            weeks_parts["weeks_list"] = list(map(int, weeks))
            return weeks_parts

        return weeks_parts

    def get_week_parity(self, lesson: str) -> int:
        """Получить четность недели"""
        for i in reversed(self.parity):
            if i in lesson:
                return self.parity.index(i) + 1
        return 0

    def get_disc_name(self, lesson: str, lesson_parts: dict) -> str:
        """Убрать лишнее из названия дисциплины и вернуть только само название"""
        disc_name = deepcopy(lesson)

        # убираем недели
        if lesson_parts["parity"] == 0 and len(lesson_parts["weeks_list"]) < 16:
            disc_name = disc_name.removeprefix(lesson_parts["weeks_text"])
        elif lesson_parts["parity"] > 0:
            parity_str = self.parity[lesson_parts["parity"] - 1]
            disc_name = disc_name.removeprefix(parity_str)

        # убираем тип пары
        if lesson_parts["worktype"] != "пр":
            disc_name = disc_name.removesuffix("лк")

        # убираем подгруппу
        if lesson_parts["sub_group"] != 0:
            sub_group_str = self.subgroups[lesson_parts["sub_group"] - 1]
            disc_name = disc_name.removesuffix(sub_group_str)

        # чтобы избежать опечаток с пробелами, добавляем после всех точек пробел

        # Лин. алг.и ан. геом. -> Лин.алг.и ан.геом.
        disc_name = disc_name.replace(". ", ".")
        # Лин.алг.и ан.геом.   -> Лин. алг. и ан. геом.
        disc_name = disc_name.replace(".", ". ")

        # убираем лишние пробелы по краям
        disc_name = disc_name.removeprefix(" ")
        disc_name = disc_name.removesuffix(" ")

        return disc_name

    def get_semcode(self, title: str) -> int:
        """Получение кода семестра по заголовку"""
        # парсим заголовок и получаем код семестра:
        # осень - 00, весна - 01

        # год в формате 2425, где 2024 - год начала учебного года, 2025 - год конца
        year_code = str(self.start_year)[-2:] + str(self.end_year)[-2:]

        semcode = "00"
        low_title = title.lower()
        is_spring = low_title.find(SPRING_SUBSTR)
        is_summer = low_title.find(SUMMER_SUBSTR)
        # если это весенний семестр, или летняя сессия, то четный семестр
        if is_spring != -1 or is_summer != -1:
            semcode = "01"
        semcode = str(year_code) + semcode

        return int(semcode)

    # TODO переименовать
    def fill_group_day_db(
            self,
            lesson_parts: dict,
            teacher: str,
            order: int,
            weekday: str,
            group_id: int,
            room: str,
            department_id: int
    ):
        """Заполнение таблиц по данным одного дня недели определенной группы"""
        # таблица дисциплин
        disc_id = self.db.set_disc(
            title="null",
            shorttitle=lesson_parts["disc_name"],
            department_id=department_id,
            varmask="null",
        )

        # таблица расписания
        # TODO глянуть с НИР у маг 2 курс, куда какой НИР пишется, с каким disc_id
        weekday_num = self.week_strs.index(weekday)
        rasp7_id = self.db.set_rasp7(
            semcode=self.semcode,
            version=self.version,
            disc_id=disc_id,
            weekday=weekday_num,
            pair=order,
            weeksarray=lesson_parts["weeks_list"],
            weekstext=lesson_parts["weeks_text"],
        )

        # таблица связи пары
        self.db.set_rasp7_groups(
            rasp7_id=rasp7_id, group_id=group_id, sub_group=lesson_parts["sub_group"]
        )

        # распределение пар по преподавателям
        # disc_teachers = []
        # if teacher.find("\n") != -1:
        #     disc_teachers = teacher.split("\n")
        # else:
        #     disc_teachers.append(teacher)

        # # если преподавателей несколько, делаем несколько записей
        # prep_ids = []
        # # for t in disc_teachers:
        # #     # таблица преподавателей
        # #     prep_ids.append(self.db.set_prep(fio=t))

        # for prep_id in prep_ids:
        #     if prep_id is not None:
        prep_id = self.db.set_prep(fio=teacher)
        self.db.set_rasp7_preps(rasp7_id=rasp7_id, prep_id=prep_id)

        # заполнение аудиторий
        self.db.set_rasp7_rooms(rasp7_id=rasp7_id, room=room)
        # заполнение таблиц rasp18
        self.fill_rasp18_for_disc(
            lesson_parts["weeks_list"],
            weekday,
            order,
            lesson_parts["worktype"],
            disc_id,
            prep_id,
            room,
            group_id,
            lesson_parts["sub_group"]
        )

    def fill_rasp18_for_disc(
            self, weeksarray, weekday, order, worktype, disc_id, prep_id, room, group_id, subgroup
    ):
        """Заполнение rasp18 для дисциплины на до конца семестра"""
        day_order = (self.week_strs.index(weekday) - 1) % 7
        first_day = datetime.strptime(
            self.start_date, "%Y-%m-%d").date()
        weekday_delta = abs(day_order - first_day.weekday())

        weekday_num = self.week_strs.index(weekday)
        params = {"semcode": self.semcode, "day": str(first_day),
                  "weekday": weekday_num, "week": 1}
        for week in weeksarray:
            # считаем дату пары
            cur_delta = weekday_delta + (week - 1) * 7
            cur_date = first_day + timedelta(days=cur_delta)
            params["week"] = week
            params["day"] = str(cur_date)

            # смотрим id дня
            day_id = self.db.get_id(table_name="sc_rasp18_days", params=params)
            # время пары
            time_start = utils.get_time_by_order(order)
            time_end = utils.time_in_90_minutes(time_start)

            rasp18_id = self.db.set_rasp18(
                semcode=self.semcode,
                day_id=day_id,
                pair=order,
                kind=0,
                worktype=worktype,
                disc_id=disc_id,
                timestart=str(time_start),
                timeend=str(time_end)
            )

            self.db.set_rasp18_groups(rasp18_id, group_id, subgroup)
            if prep_id is not None:
                self.db.set_rasp18_preps(rasp18_id, prep_id)
            if room is not None:
                self.db.set_rasp18_rooms(rasp18_id, room)

    def fill_rasp_title_parts(self, max_row: int, ws):
        """
            Заполнить по заголовоку расписания нужные для парсера данные:
            учебные года, версия, время года семестра

            return start_year, end_year, version
        """
        rasp_title = ""
        is_version = False
        for row in range(1, max_row):
            rasp_title = ws.cell(row, self.version_col).value
            if rasp_title is None or not isinstance(rasp_title, str):
                continue

            # если версия есть
            if rasp_title.find("версия") != -1:
                is_version = True
                break

        if not is_version:
            raise ValueError("Версия расписания не была найдена")

        self.version = utils.get_version(rasp_title)
        self.start_year, self.end_year = utils.get_stud_years(rasp_title)
        self.semcode = self.get_semcode(rasp_title)
        # TODO подумать и мб надо такое сделать
        # self.start_date, self.end_date = utils.get_stud_period(self.semcode)

    def get_disc_id(self, lesson_name: str, department_id: int) -> int:
        """Получить id дисциплины по названию"""
        params = {"shorttitle": lesson_name, "department_id": department_id}
        disc_id = self.db.get_id(table_name="sc_disc", params=params)
        return disc_id

    def get_dep_id(self, cel_color: int) -> int:
        """Получить по цвету ячейки id кафедры"""
        department_name = "другая"
        match cel_color:
            # зеленый
            case "FFCCFF66":    # только для ВЕГИ
                department_name = "только для ВЕГИ"
            # розовый
            case "FFFFCCFF":    # только для ВМ
                department_name = "только для ВМ"
            # темно желтый
            case "FFFFE15A":    # Для всех, ведет кафедра ВМ
                department_name = "ВМ"
            case "FFFFF56D":    # Для всех, ведет кафедра ВМ/Просто для всех(у бакалавров)
                department_name = "ВМ"
            # желтый/бледно желтый
            case "FFF1FF67":    # 
                department_name = "ВЕГА"
            case "FFF4FF67":    # ведет ВЕГА для всех
                department_name = "ВЕГА"
            case "FFEAFF9F":    # ведет ВЕГА
                department_name = "ВЕГА"
            case "FFE5FF99":    # ведет ВЕГА
                department_name = "ВЕГА"
            # голубой цвет
            case "FFB2ECFF":    # Для всех, ведут другие кафедры
                department_name = "другая"
            case "FFD1F3FF":    # Для всех, ведут другие кафедры
                department_name = "другая"
            # белый/без заливки
            case "0":
                department_name = "другая"
            case "00000000":
                department_name = "пустая"

        dep_params = {"title": department_name}
        department_id = self.db.get_id("sc_department", dep_params)
        return department_id


if __name__ == "__main__":
    parser = VegaRaspParser()
    parser.parse()
