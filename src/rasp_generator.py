"""Генерация расписания в xlsx формате"""

from configparser import ConfigParser

from enum import IntEnum

from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill

from .cell_colors import CellColors
from .db_class import Database


class HelpLists:
    """Списки с вспомогательными данными"""
    weekdays_names = ["ВС", "ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ"]


class RaspConsts(IntEnum):
    """Позиции и другие константы расписания"""
    GROUP_ROW = 3
    WEEKDAYS = 6
    PAIR_PER_DAY = 7
    PAIR_PER_CELL = 2

    WEEKDAY_MERGE = PAIR_PER_DAY*2-1


class RaspGenerator:
    """Класс генератора расписания, по данным загруженным в БД"""

    def __init__(self):
        """Инициализация генератора"""
        config = ConfigParser()
        config.read("config.ini")
        if not config.has_option("RASP_PARAMS", "semcode") or \
                not config.has_option("RASP_PARAMS", "version") or \
                not config.has_option("RASP_PARAMS", "version_date"):
            raise ValueError("Отсутствую параметры или config.ini для начала работы генератора")

        self.semcode = config["RASP_PARAMS"]["semcode"]
        self.version = config["RASP_PARAMS"]["version"]
        self.version_date = config["RASP_PARAMS"]["version_date"]

        self.order_fill = PatternFill(patternType='solid', start_color=CellColors.ORDER)
        self.weekday_fill = PatternFill(patternType='solid', start_color=CellColors.BORDER)
        self.db_conn = Database()

    def generate_rasp(self):
        """Сгенерировать расписание"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Занятия"

        self.create_rasp_title(ws)
        self.fill_day_col(ws=ws, col_num=1, is_left=True)
        groups_rasp = self.get_week_groups_rasp()
        for g_rasp in groups_rasp:
            self.fill_group_col(g_rasp)
        # self.fill_day_col(ws=ws, col_num=4, is_left=False)
        wb.save("new_rasp.xlsx")

    def create_rasp_title(self, ws):
        """Заполняем шапку"""
        semcode_data = self.get_semcode_data()
        rasp_title = f"""РАСПИСАНИЕ  ЗАНЯТИЙ  НА  {semcode_data["season"]}  СЕМЕСТР
 {semcode_data["years"]}  УЧЕБНОГО  ГОДА\nверсия {self.version} от {self.version_date}"""
        rasp_title.replace("\n ", " ")
        ws.cell(1, 1).value = rasp_title

    def get_semcode_data(self):
        """Сезон года по коду семестра"""
        semcode_data = {"season": "ОСЕННИЙ", "years": "2024/25"}
        season_code = str(self.semcode)[:-2]
        if season_code == "01":
            semcode_data["season"] = "ВЕСЕННИЙ"
        semcode_years = str(self.semcode)[:4]
        semcode_data["years"] = f"20{semcode_years[:2]}/{semcode_years[2:]}"
        return semcode_data

    def fill_day_col(self, ws, col_num: int, is_left: bool):
        """
            Заполняет столбец днями недели и номерами пар(занимает два столбца)
            is_left - столбец на левой границе расписания или на правой
        """
        weekday_col = col_num
        order_col = col_num + 1
        if not is_left:
            weekday_col, order_col = order_col, weekday_col

        last_row = RaspConsts.PAIR_PER_DAY * RaspConsts.PAIR_PER_CELL * RaspConsts.WEEKDAYS
        last_row += RaspConsts.GROUP_ROW

        cur_order = 1
        cur_weekday_idx = 1
        for row in range(RaspConsts.GROUP_ROW + 1, last_row, 2):
            # если начало дня
            if cur_order == 1:
                ws.cell(row, weekday_col).value = HelpLists.weekdays_names[cur_weekday_idx%7]
                ws.cell(row, weekday_col).fill = self.weekday_fill
                ws.merge_cells(start_row=row, start_column=weekday_col,
                               end_row=row+RaspConsts.WEEKDAY_MERGE, end_column=weekday_col)

            ws.cell(row, order_col).value = cur_order
            ws.cell(row, order_col).fill = self.order_fill
            ws.merge_cells(start_row=row, start_column=order_col,
                           end_row=row+1, end_column=order_col)
            cur_order += 1

            # если конец дня -> новый день недели
            if cur_order == 8:
                cur_order = 1
                cur_weekday_idx += 1

    def get_week_groups_rasp(self):
        """Получить раписание для всех групп обоих подгрупп на неделю"""
        group_data = self.db_conn.get_groups_data()
        print(group_data)
        group_rasp = {}
        for group in group_data:
            g_rasp = self.db_conn.get_discs_for_group(self.semcode, group["id"])
            group_rasp[group["title"]] = g_rasp
        return group_rasp

    def fill_group_col(self, group_rasp):
        """Заполняем расписание для группы на неделю"""
        # если у выбранной группы нет пар в выбранном семестре
        if group_rasp is None:
            return
        
        
